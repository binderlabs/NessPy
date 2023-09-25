#!/usr/bin/python3

import requests
import argparse
from termcolor import colored
import json
import sys
import time
import os
import progressbar
from bs4 import BeautifulSoup
import xlsxwriter
import openpyxl as xl
from datetime import date
import re

logo = """
     _   _               ____        
    | \ | | ___  ___ ___|  _ \ _   _ 
    |  \| |/ _ \/ __/ __| |_) | | | |
    | |\  |  __/\__ \__ \  __/| |_| |
    |_| \_|\___||___/___/_|    \__, |
                               |___/"""


url = "https://localhost:8834"
username= "admin"
password= "password"
path = "output"
compliance_path = "compliance"
previous_audit_path = "previous_audit"



# Don't verify the SSL certificate
verify = False

# Disable warning when not verifying SSL certificates
requests.packages.urllib3.disable_warnings(requests.packages.urllib3.exceptions.InsecureRequestWarning)


usage = "nesspy [-t TARGET | -T TARGET FILE] -p [CUSTOM POLICY NAME] -f [FOLDER NAME] -e [EXPORT FORMAT] -n [SCAN NAME] -o [OUTPUT FILENAME WITHOUT EXTENSION] [ -c | --compliance] \n" + logo
example="""Examples:
       nesspy -l
       nesspy -t 127.0.0.1 -p 'My Custom Policy' -n 'My first Nessus Scan' -e csv -o 127.0.0.1
       nesspy -t 127.0.0.1 -p 'My Custom Policy' -e csv,html,nessus -o 127.0.0.1
       nesspy -T list.txt -p 'My Custom Policy' -f 'NESSUS FOLDER' -n 'Production Machine list' -o list
       nesspy -t 127.0.0.1 -p 'My Compliance Policy' -f 'compliance folder' -e 'csv,html,nessus' -o 'compliance_assessment_127.0.0.1' --compliance
"""

parser = argparse.ArgumentParser(usage=usage,epilog=example,formatter_class=argparse.RawDescriptionHelpFormatter)
group = parser.add_mutually_exclusive_group()
parser.add_argument('-l','--list-policies',dest="list_policies",action="store_true",help="Lists all policies.")
group.add_argument('-t',dest="target",help="Single target to launch scan against.")
group.add_argument('-T',dest="target_file",help="File with a list of targets.")
parser.add_argument("-p",dest='policy_name',type=str,help='Custom policy to use for nessus scan.')
parser.add_argument("-f",dest='folder_name',type=str,help='Folder to store nessus scan.')
parser.add_argument("-n  ",dest='scan_name',type=str,help='Name to be used for the particular scan. If not specified, default value will be the target name.')
parser.add_argument('-e  ',dest="export_format",help="Export the scan report in specified format.\nThe available formats are nessus,html,csv and db.")
parser.add_argument('-o  ',dest="output",type=str,help="File to output the result.")
parser.add_argument('-c','--compliance',action="store_true",help="Compliance scan")
args = parser.parse_args()

if not len(sys.argv)>1:
    parser.print_help()
    exit()

def clean():
    
# For macOS and Linux
    if os.name == 'posix':
        os.system('clear')


    # Windows
    elif os.name == 'nt':
        os.system('cls')



def login():
    r = requests.post(url + '/session',data={'username':username,'password':password},verify=verify)
    if(r.status_code==200):
        global token
        token = r.json().get('token')
    else:
        print("[!] " + r.json()['error'])
        exit()

def logout():
    requests.delete(url + '/session',headers=headers,verify=verify)
    print(colored("[*] Session killed.",'green',attrs=['bold']))
    exit()

def getApiToken():
    res = requests.get(url +"/nessus6.js",verify=False)
    result = re.findall(r"key:\"getApiToken\",value:function\(\){return\".*",res.text)
    string = result[0]
    api_token = ""
    for i in range(42,78):
        api_token += string[i]
    return api_token


def list_policies():

    policies = requests.get(url + '/policies',headers=headers,verify=verify)   # Custom Policies
    templates = requests.get(url+"/editor/policy/templates",headers=headers,verify=False)       
    if policies.status_code==403 or templates.status_code==403:
        print(colored("[!] User don't have the permission to view the policy list",'yellow',attrs=['bold']))
    if policies.status_code==200 and templates.status_code==200:
        return policies.json()['policies'],templates.json()['templates']        
    else:
        print("[!] " + str(policies.json()['error']))
        logout()


def policy_json_from_policy_name(): 
    policies = list_policies()                                 
    for policy in policies[0]:                  
        if args.policy_name==policy['name']:
            return policy
    for policy in policies[1]:                         
        if args.policy_name==policy['title']:         
            return policy
    print(colored("[!] Cannot find the policy with name " + args.policy_name,'red',attrs=['bold']))
    logout()      

def create_scan(uuid,name,targets,id):
    print("[*] Creating scan")

    # Check folders
    if args.folder_name:
        # Create folder
        
        folder = requests.post(url + "/folders",json={'name':args.folder_name},verify=verify,headers=headers)

        if folder.status_code == 200:
            folder_id = folder.json()['id']
            print("[*] Folder created with id " + str(folder_id)+ ".\n" )
            payload = {"uuid" : uuid,
                "settings" :  {
                    "name" : name,
                    "text_targets" : targets,
                    "folder_id" : folder_id,
                    "launch_now" : True,
                               }
               }
            if not id=="":              
                payload["settings"]["policy_id"] = id

        elif str(folder.json()['error']) == "A folder with the same name already exists":
            folders = requests.get(url + "/folders",verify=verify,headers=headers).json()

            for folder in folders['folders']:
                if args.folder_name == folder['name']:
                    folder_id = folder['id']
            
            payload = {"uuid" : uuid,
                "settings" :  {
                    "name" : name,
                    "text_targets" : targets,
                    "folder_id" : folder_id,
                    "launch_now" : True,
                               }
               }
            if not id=="":              
                payload["settings"]["policy_id"] = id
            
            
            
        else:
            print("[!] " + str(folder.json()['error']))
            logout()




    else:
        # Store the scan result in default My Scans folder.
        payload = {"uuid" : uuid,
                "settings" :  {
                    "name" : name,
                    "text_targets" : targets,
                    "launch_now" : True,
                               }
               }
        if not id=="":               # Used when policy id is supplied
            payload["settings"]["policy_id"] = id

    payload = json.dumps(payload)
    new_scan = requests.post(url + '/scans',data=payload,headers=headers,verify=verify)

    if new_scan.status_code==200:
        print(colored("[*] Scan created successfully.",'green',attrs=['bold']))
        print("[*] Launching scan")
        return new_scan.json()['scan']['id']
    else:
        print(str(new_scan.json()['error']))
        logout()


def show_status(scan_id):
    print("\n" + colored('[*] Status : ','blue',attrs=['bold']) + colored('Running','yellow',attrs=['bold']))
    bar = progressbar.ProgressBar(100)
    bar.start()
    while True:
        status = requests.get(url + '/scans/' + str(scan_id),headers=headers,verify=verify)
        status = status.json()['info']['status']
        if status=='canceled':
            clean()
            print(colored("[!] Scan has been Cancelled.",'red',attrs=['bold']))
            logout()

        if status=='paused':
            clean()
            print(colored("[!] Scan has been Paused.",'yellow',attrs=['bold']))
            logout()

        if status=='completed':
            clean()
            print(colored("[*] Scan Completed!\n",'green',attrs=['bold']))
            break

        if status=='running':
            try:
                current_progress = requests.get(url + '/scans/' + str(scan_id),headers=headers,verify=verify)

                current_progress = current_progress.json()['hosts'][0]['scanprogresscurrent']
                for i in range(100):
                    i = current_progress
                    bar.update(i)
            except:
                continue

            time.sleep(2)


def export_request(scan_id):
    print("[*] Exporting report(s)")

    format_list = args.export_format.split(',')
    for export_format in format_list:
        if export_format =='csv' or export_format =='nessus':
            payload =  { "format" : export_format }

        elif export_format =='db':
            payload =  { "format" : export_format , "password":password}

        elif export_format == 'pdf' or export_format =='html':
            if args.compliance:
                res = requests.get(url + '/reports/custom/templates',headers=headers,verify=verify)
                id = res.json()[0]['id']
                payload = { "format":export_format, "template_id":int(id)} # Template: Detailed Vulnerabilities by Host with Compliance/Remediations
                
            else:
                payload = { "format":export_format, "chapters":"vuln_hosts_summary"}

        else:
            print(colored("[!] Unsupported format detected!",'red',attrs=['bold']))
            logout()

        payload = json.dumps(payload)
        res = requests.post(url + '/scans/' + str(scan_id) + '/export',data=payload,verify=verify,headers=headers)

        if res.status_code==200:
            file_id = res.json()['file']
            print("[*] Report Generating")
            time.sleep(2)
            while export_status(scan_id,file_id) is False:
                time.sleep(1)
            export_download(scan_id,file_id,export_format)

        else:
            print("[!] " + res.json()['error'])
            print("[!] Waiting for 10 seconds before retrying...")
            time.sleep(10)
            export_request(scan_id)

    
def export_status(scan_id,file_id):
    res = requests.get(url + '/scans/{0}/export/{1}/status'.format(scan_id,file_id),headers=headers,verify=verify)
    return res.json()['status']=='ready'
    


def export_download(scan_id,file_id,export_format):
        
    res = requests.get(url + '/scans/' + str(scan_id) + '/export/' + str(file_id) +'/download',headers=headers,verify=verify)
    
    if res.status_code != 200:
        print("[!] " + res.json()['error'])
        export_download(scan_id,file_id,export_format)

    else:
        if args.compliance and export_format == 'html':       
            write_compliance_excel(res.content)

        print(colored("[*] Report is ready to download!",'green',attrs=['bold']))
        print("[*] Downloading the report")                
        print("[*] Report downloaded")
        print("[*] Storing the report downloaded")
        if args.output:
            filename = args.output + "." + export_format
        else:
            filename = 'nessus_{0}_{1}.{2}'.format(scan_id,file_id,export_format)

        
        path = "output"
        isExist = os.path.exists(path)
        if not isExist:
            os.makedirs(path)
        
        with open(path + "/" + filename,'wb') as f:
            f.write(res.content)
        print(colored("[*] Output stored to " + path + "/" + filename + "\n",'green',attrs=['bold']))

def write_compliance_excel(response):
    soup = BeautifulSoup(response,'html.parser')
    tags = soup.find_all("ul")
    contents = tags[2] # Failed Compliances
    failed_compliances = contents.find_all("li")

    if failed_compliances != []:
        print("[*] Saving compliance excel files.")  
        compliance_path = "compliance"
        isExist = os.path.exists(compliance_path)
        if not isExist:
            os.makedirs(compliance_path)

        today = date.today()
        filename = compliance_path + "/" + str(args.target)+"-Compliance Assessment-"+ str(today.strftime('%Y-%b'))+ ".xlsx"
        wb = xlsxwriter.Workbook(filename)
        f1 = wb.add_format()
        f1.set_bold()
        ws = wb.add_worksheet(args.target)
    
        row = 1

        for failed_compliance in failed_compliances:
            ws.write(row,0,failed_compliance.text)
            row += 1
        table = "A1:F" + str(row)

        title = "Compliance Assessment: " + args.target
        ws.add_table(table, {'columns':[{'header':title},{'header':'Remarks'},{'header':'Justification'},{'header':'Previous Audit BDO Comments'},{'header':'BDO Comments'},{'header':'BDO Status'},],'style': None})
        ws.set_row(0,None,f1)
        ws.autofit()
        wb.close()
        print(colored("[*] Output stored to " + filename + "\n",'green',attrs=['bold']))
        invoke_vlookup(filename)

    
    else:
        # Need to check whether host is alive manually when this message is shown.
        print(colored("[*] No failed compliance cases detected for "+ args.target + "\n",'green',attrs=['bold']))
    

def invoke_vlookup(filename):
    filename = filename
    isExist = os.path.exists(previous_audit_path)
    if not isExist:
        print(colored("[!] Previous compliance audit directory not found!",'yellow',attrs=['bold']))
    


    else:
        files = os.listdir(previous_audit_path)
        if files != []:
            cwd = os.getcwd()
            for file in files:
                # Check IP address in filename
                if args.target in file:
                    # Assume only have 1 worksheet since the filename is named with IP address.
                    wb = xl.load_workbook('./'+ filename)

                    sheet = wb.active
                    max_row = sheet.max_row
                    
                    for row in range(1,max_row):
                        cell = sheet.cell(row=row+1, column=3)
                        if os.name == 'posix':
                            justification_vlookup = "=VLOOKUP(A"+str(row+1)+",'["+str(cwd)+"/"+previous_audit_path+"/"+file+"]"+args.target+"'!A2:F500,3,0)"
                        elif os.name == 'nt':
                            justification_vlookup = "=VLOOKUP(A"+str(row+1)+",'"+str(cwd)+"\\"+previous_audit_path+"\\["+file+"]"+args.target+"'!A2:F500,3,0)"
                        cell.value = justification_vlookup

                        cell = sheet.cell(row=row+1, column=4)
                        if os.name == 'posix':
                            previous_comment_vlookup = "=VLOOKUP(A"+str(row+1)+",'["+str(cwd)+"/"+previous_audit_path+"/"+file+"]"+args.target+"'!A2:F500,4,0)"
                        elif os.name == 'nt':
                            previous_comment_vlookup = "=VLOOKUP(A"+str(row+1)+",'"+str(cwd)+"\\"+previous_audit_path+"\\["+file+"]"+args.target+"'!A2:F500,4,0)"

                        cell.value = previous_comment_vlookup


                    wb.save('./'+filename)
                    break
                
                elif '.xlsx' in file:
                    wb = xl.load_workbook('./'+ previous_audit_path + "/" + file)
                    sheetnames = wb.sheetnames

                    for sheetname in sheetnames:
                        if args.target in sheetname:
                            wb = xl.load_workbook('./'+ filename)
                            sheet = wb[sheetname]
                            max_row = sheet.max_row
                    
                            for row in range(1,max_row):

                                cell = sheet.cell(row=row+1, column=3)
                                if os.name == 'posix':
                                    justification_vlookup = "=VLOOKUP(A"+str(row+1)+",'["+str(cwd)+"/"+previous_audit_path+"/"+file+"]"+args.target+"'!A2:F500,3,0)"
                                elif os.name == 'nt':
                                    justification_vlookup = "=VLOOKUP(A"+str(row+1)+",'"+str(cwd)+"\\"+previous_audit_path+"\\["+file+"]"+args.target+"'!A2:F500,3,0)"
                                cell.value = justification_vlookup

                                cell = sheet.cell(row=row+1, column=4)
                                if os.name == 'posix':    
                                    previous_comment_vlookup = "=VLOOKUP(A"+str(row+1)+",'["+str(cwd)+"/"+previous_audit_path+"/"+file+"]"+args.target+"'!A2:F500,4,0)"
                                elif os.name == 'nt':
                                    previous_comment_vlookup = "=VLOOKUP(A"+str(row+1)+",'"+str(cwd)+"\\"+previous_audit_path+"\\["+file+"]"+args.target+"'!A2:F500,4,0)"
                                cell.value = previous_comment_vlookup
                            wb.save('./'+filename)
                            break
                       
        
        else:
            print(colored("[!] Previous compliance audit directory not found!",'yellow',attrs=['bold']))


if __name__=='__main__':
    print(logo)
    
    token= ""
    api_token = getApiToken()
    
    login()
    
    headers = {'X-Cookie': 'token=' + token,'X-Api-Token': api_token , 'content-type': 'application/json'}  

    # For debugging purposes. 
    proxies = {
   'http': 'http://127.0.0.1:8080',
   'https': 'http://127.0.0.1:8080',
   }

    
    if args.list_policies:
        policies = list_policies()
        if not policies[0]==None:
            print("\n\n")
            print("--------------------------")
            print("   CUSTOM POLICIES  ")
            print("--------------------------\n")
            for policy in policies[0]:
                print(policy['name'])


        print("\n\n" + "--------------------------")
        print("   SCAN TEMPLATES  ")
        print("--------------------------\n")
        for policy in policies[1]:
            print(policy['title'])
    
   
    if args.compliance and args.target == None:
        print(colored("\n"+ "[!] Please specify a single target with -t for compliance scan.",'red',attrs=['bold']))
        exit(-1)

    if args.target:
        target=args.target

    if args.target_file:
        try:
            with open(args.target_file) as t:
                target=t.read()
        except IOError:
            print(colored("[!] Error opening the file: " + args.target_file,'red',attrs=['bold']))
            exit()


    # If target and policy name both are specified, launch the scan.
    if args.policy_name and (args.target or args.target_file):

        if args.target:
            name = args.target

        elif args.target_file:
            name = args.target_file

        if args.scan_name:
            name = args.scan_name
        policy = policy_json_from_policy_name()

    
        try:
            scan_id = create_scan(policy['template_uuid'],name,target,policy['id'])
        except KeyError:
            scan_id = create_scan(policy['uuid'],name,target,"")
        show_status(scan_id)

        # Try to export the scan report in specified format, when -e flag is set
        if args.export_format:
            export_request(scan_id)

    logout()

    
